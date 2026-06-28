import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def generate_report(purchases, date_from, date_to) -> bytes:
    """
    OUT-REP: donor report with header, table, and footer totals.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Звіт"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Звіт БФ «Сила Єдності» за період {date_from} — {date_to}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")

    headers = ["Дата закупівлі", "Підрозділ", "Що передано", "Категорія", "Сума (грн)"]
    header_fill = PatternFill(fill_type="solid", fgColor="1F497D")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 22

    purchases_list = list(purchases)
    total = 0
    count = 0

    for row_idx, purchase in enumerate(purchases_list, start=4):
        fill = (
            PatternFill(fill_type="solid", fgColor="F8F9FA")
            if row_idx % 2 == 0
            else None
        )
        values = [
            purchase.purchase_date,
            purchase.request.unit_name,
            purchase.request.item_name[:100],
            purchase.request.category.name,
            float(purchase.actual_cost),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 3))
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18
        total += purchase.actual_cost
        count += 1

    footer_row_1 = len(purchases_list) + 4
    footer_row_2 = footer_row_1 + 1

    ws.merge_cells(f"A{footer_row_1}:E{footer_row_1}")

    ws.cell(row=footer_row_2, column=3, value="Всього заявок виконано:").font = Font(
        bold=True, size=11
    )
    ws.cell(row=footer_row_2, column=3).alignment = Alignment(horizontal="right")
    count_cell = ws.cell(row=footer_row_2, column=4, value=count)
    count_cell.font = Font(bold=True, size=11, color="1F497D")
    count_cell.alignment = Alignment(horizontal="center")

    footer_row_3 = footer_row_2 + 1
    ws.cell(row=footer_row_3, column=3, value="Загальна сума витрат:").font = Font(
        bold=True, size=11
    )
    ws.cell(row=footer_row_3, column=3).alignment = Alignment(horizontal="right")
    sum_cell = ws.cell(row=footer_row_3, column=4, value=float(total))
    sum_cell.font = Font(bold=True, size=12, color="198754")
    sum_cell.alignment = Alignment(horizontal="center")
    ws.cell(row=footer_row_3, column=5, value="грн").font = Font(bold=True)

    for col, width in zip(["A", "B", "C", "D", "E"], [16, 28, 50, 22, 14]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
